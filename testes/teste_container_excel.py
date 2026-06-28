from core.login import login_whatsapp
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from datetime import datetime
import time

print("=" * 60)
print("🧪 TESTE: VARRENDO DIVS E SALVANDO NO EXCEL")
print("📌 Vai mostrar todas as divs, classes e textos")
print("=" * 60)

driver = login_whatsapp()

if driver:
    print("\n✅ LOGIN REALIZADO!")
    time.sleep(3)

    # Entra na primeira conversa
    print("📂 Entrando na primeira conversa...")
    conversas = driver.find_elements(By.XPATH, "//div[@role='row']")
    if conversas:
        conversas[0].click()
        time.sleep(3)
        print("✅ Dentro da conversa!")
    else:
        print("❌ Nenhuma conversa encontrada.")
        driver.quit()
        exit()

    # Varrer todas as divs
    divs = driver.find_elements(By.TAG_NAME, "div")
    print(f"\n📦 Total de divs encontradas: {len(divs)}\n")

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Divs e Classes"

    cabecalhos = ["POSIÇÃO", "TAG", "CLASSES", "TEXTO"]
    for col, titulo in enumerate(cabecalhos, 1):
        ws.cell(row=1, column=col, value=titulo)

    linha_atual = 2

    for i, div in enumerate(divs, 1):
        try:
            classe = div.get_attribute("class") or "Sem classe"
            texto = div.text.strip()[:200] if div.text else ""

            ws.cell(row=linha_atual, column=1, value=i)
            ws.cell(row=linha_atual, column=2, value="div")
            ws.cell(row=linha_atual, column=3, value=classe)
            ws.cell(row=linha_atual, column=4, value=texto)
            linha_atual += 1

        except Exception as e:
            print(f"   ⚠️ Erro na div {i}: {e}")

    nome_arquivo = f"mapa_divs_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 100
    ws.column_dimensions["D"].width = 80
    wb.save(nome_arquivo)

    print(f"   ✅ Mapa salvo: {nome_arquivo}")
    driver.quit()
else:
    print("❌ Falha no login.")
