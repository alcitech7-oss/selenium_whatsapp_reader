from selenium.webdriver.common.by import By
from openpyxl import Workbook
from datetime import datetime


def mapear_spans(driver):
    print("   🗺️  Mapeando spans por posição e estrutura...")

    spans = driver.find_elements(By.XPATH, "//span[text()]")
    print(f"   📝 Total de spans: {len(spans)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de Spans"

    cabecalhos = ["POSIÇÃO", "TEXTO", "TAG", "CLASSES", "DATA-TESTID"]
    for col, titulo in enumerate(cabecalhos, 1):
        ws.cell(row=1, column=col, value=titulo)

    linha_atual = 2

    for i, span in enumerate(spans, 1):
        try:
            texto = span.text.strip()
            if not texto:
                continue

            classes = span.get_attribute("class") or ""
            data_testid = span.get_attribute("data-testid") or ""
            pai = span.find_element(By.XPATH, "..")
            tag_pai = pai.tag_name

            posicao = f"Span {i} - dentro de <{tag_pai}>"

            ws.cell(row=linha_atual, column=1, value=posicao)
            ws.cell(row=linha_atual, column=2, value=texto)
            ws.cell(row=linha_atual, column=3, value="span")
            ws.cell(row=linha_atual, column=4, value=classes)
            ws.cell(row=linha_atual, column=5, value=data_testid)
            linha_atual += 1

        except Exception as e:
            print(f"   ⚠️ Erro no span {i}: {e}")

    nome_arquivo = f"mapa_spans_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(nome_arquivo)
    print(f"   ✅ Mapa salvo: {nome_arquivo}")
    return nome_arquivo
