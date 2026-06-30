from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from selenium.webdriver.common.by import By
from core.selectors import CONTAINER_CONVERSA, REMETENTE, MENSAGEM, HORARIO, NAO_LIDAS
from datetime import datetime
import os
import spacy
from langdetect import detect, DetectorFactory
from langdetect.lang_detect_exception import LangDetectException

# Garante reprodutibilidade na detecção de idioma
DetectorFactory.seed = 0

# Carrega modelos spaCy (fallback se não encontrar)
try:
    nlp_pt = spacy.load("pt_core_news_sm")
except:
    print(
        "   ⚠️ Modelo pt_core_news_sm não encontrado. Instale com: python -m spacy download pt_core_news_sm"
    )
    nlp_pt = None

try:
    nlp_en = spacy.load("en_core_web_sm")
except:
    print(
        "   ⚠️ Modelo en_core_web_sm não encontrado. Instale com: python -m spacy download en_core_web_sm"
    )
    nlp_en = None


def detectar_idioma(texto):
    """Detecta o idioma da mensagem (pt ou en)"""
    try:
        idioma = detect(texto)
        if idioma in ["pt", "pt-br"]:
            return "pt"
        elif idioma == "en":
            return "en"
        else:
            return "pt"  # fallback para português
    except LangDetectException:
        return "pt"  # fallback se não conseguir detectar


def classificar_mensagem(texto, idioma):
    """Classifica a mensagem com base no idioma detectado"""
    if idioma == "pt" and nlp_pt:
        doc = nlp_pt(texto.lower())
    elif idioma == "en" and nlp_en:
        doc = nlp_en(texto.lower())
    else:
        # Fallback: usa o modelo que estiver disponível
        doc = nlp_pt(texto.lower()) if nlp_pt else None

    if doc is None:
        return "Outros"

    lemas = [token.lemma_ for token in doc]

    # Palavras-chave por idioma e categoria
    if idioma == "pt":
        palavras = {
            "Pedidos": [
                "querer",
                "pedir",
                "comprar",
                "encomendar",
                "gostar",
                "precisar",
            ],
            "Reclamações": [
                "problema",
                "defeito",
                "reclamar",
                "erro",
                "ruim",
                "falha",
                "queimar",
                "atrasar",
            ],
            "Entregas": [
                "entregar",
                "motoboy",
                "retirar",
                "endereço",
                "chegar",
                "sair",
            ],
            "Fornecedores": [
                "orçamento",
                "estoque",
                "preço",
                "fornecedor",
                "lote",
                "reposição",
            ],
            "Saudações": ["oi", "olá", "bom", "tudo", "beleza"],
        }
    else:  # inglês
        palavras = {
            "Pedidos": ["want", "order", "buy", "get", "need", "would like"],
            "Reclamações": [
                "problem",
                "defect",
                "complain",
                "error",
                "bad",
                "fail",
                "broken",
                "late",
            ],
            "Entregas": [
                "deliver",
                "delivery",
                "motoboy",
                "pickup",
                "address",
                "arrive",
                "leave",
            ],
            "Fornecedores": [
                "budget",
                "stock",
                "price",
                "supplier",
                "batch",
                "replenish",
            ],
            "Saudações": ["hi", "hello", "good", "how are you", "hey"],
        }

    for categoria, palavras_chave in palavras.items():
        if any(p in lemas for p in palavras_chave):
            return categoria

    return "Outros"


def extract_containers(driver, filename):
    print(f"   📊 Extracting containers... (file: {filename})")

    if not os.path.exists(filename):
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        abas = [
            "Pedidos",
            "Reclamações",
            "Entregas",
            "Fornecedores",
            "Saudações",
            "Outros",
        ]
        for aba in abas:
            ws = wb.create_sheet(aba)
            headers = ["SENDER", "MESSAGE", "TIME", "UNREAD", "EXTRACTION_TIMESTAMP"]
            for col, title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=title)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(
                    start_color="4F81BD", end_color="4F81BD", fill_type="solid"
                )
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 25
    else:
        wb = load_workbook(filename)

    last_messages = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0] and row[1]:
                    last_messages[(sheet_name, row[0])] = row[1]

    extraction_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_added = 0

    try:
        containers = driver.find_elements(By.XPATH, CONTAINER_CONVERSA)
        print(f"   📦 Containers found: {len(containers)}")

        for i, container in enumerate(containers, 1):
            try:
                sender = container.find_element(By.XPATH, REMETENTE).text
            except:
                sender = "UNKNOWN"
            try:
                message = container.find_element(By.XPATH, MENSAGEM).text
            except:
                message = ""
            try:
                time = container.find_element(By.XPATH, HORARIO).text
            except:
                time = ""
            try:
                unread = container.find_element(By.XPATH, NAO_LIDAS).text
            except:
                unread = "0"

            if not message:
                continue

            # Detecta idioma e classifica
            idioma = detectar_idioma(message)
            target_sheet = classificar_mensagem(message, idioma)

            key = (target_sheet, sender)
            if key in last_messages and last_messages[key] == message:
                print(f"   ⏭️  Container {i}: duplicate message (ignored)")
                continue

            ws_target = wb[target_sheet]
            next_row = ws_target.max_row + 1
            ws_target.cell(row=next_row, column=1, value=sender)
            ws_target.cell(row=next_row, column=2, value=message)
            ws_target.cell(row=next_row, column=3, value=time)
            ws_target.cell(row=next_row, column=4, value=unread)
            ws_target.cell(row=next_row, column=5, value=extraction_timestamp)
            total_added += 1
            last_messages[key] = message
            print(
                f"   ✅ Container {i}: {sender} - {message[:30]}... (NEW in {target_sheet} | {idioma})"
            )

    except Exception as e:
        print(f"   ❌ Error extracting: {e}")

    try:
        wb.save(filename)
        print(f"   ✅ Spreadsheet updated: {filename}")
    except PermissionError:
        print("\n" + "=" * 60)
        print("⚠️  FILE OPEN DETECTED!")
        print(f"📌 Close the file '{filename}' in Excel.")
        print("=" * 60 + "\n")
    except Exception as e:
        print(f"   ❌ Error saving: {e}")

    print(f"   📊 Total added: {total_added}")
    return filename
