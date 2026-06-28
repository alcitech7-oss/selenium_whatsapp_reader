from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from selenium.webdriver.common.by import By
from core.seletores_classe import REMETENTE, MENSAGEM
from datetime import datetime
import os


def escrever_master(driver, nome_arquivo=None):
    if nome_arquivo is None:
        hoje = datetime.now().strftime("%Y-%m-%d")
        nome_arquivo = f"mensagens_{hoje}.xlsx"

    print(f"   📊 Escrevendo com base no mapa... (arquivo: {nome_arquivo})")

    if os.path.exists(nome_arquivo):
        wb = load_workbook(nome_arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mensagens e Remetentes"

        cabecalhos = ["REMETENTE", "MENSAGEM"]
        for col, titulo in enumerate(cabecalhos, 1):
            celula = ws.cell(row=1, column=col, value=titulo)
            celula.font = Font(bold=True, color="FFFFFF")
            celula.alignment = Alignment(horizontal="center", vertical="center")
            celula.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60

    try:
        remetente = driver.find_element(By.XPATH, REMETENTE).text
        mensagem = driver.find_element(By.XPATH, MENSAGEM).text

        if remetente and mensagem:
            ws.cell(row=ws.max_row + 1, column=1, value=remetente)
            ws.cell(row=ws.max_row, column=2, value=mensagem)
            print(f"   ✅ Salvo: {remetente} - {mensagem[:50]}...")
    except Exception as e:
        print(f"   ❌ Erro: {e}")

    wb.save(nome_arquivo)
    print(f"   ✅ Planilha salva: {nome_arquivo}")
    return nome_arquivo
